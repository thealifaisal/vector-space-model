import math
from datetime import datetime

import nltk
import openpyxl

import operator
from src.pre_processing import Preprocessing


class VSM:

    def import_stop_list(self, path):
        # a stop-word file is opened and parsed to be saved as a list
        stop_file = open(path, "r")
        stop_list = stop_file.read().split("\n")
        stop_file.close()
        return stop_list

    def process_query(self, _query):

        print(datetime.now().strftime("%H:%M:%S") + ": initiating pre-processing of query...")

        # lemma_set = {lemma-1: tf, lemma-2: tf, ...}
        lemma_set = self.pre_processing("../resource/stopword-list.txt", _query)

        print(datetime.now().strftime("%H:%M:%S") + ": query pre-processing completed")
        print(datetime.now().strftime("%H:%M:%S") + ": query lemma-set > " + str(lemma_set))

        return lemma_set

    def calculate_query_score(self, _doc_sheet, _lemma_set):

        print(datetime.now().strftime("%H:%M:%S") + ": initiating query tf-idf calculations...")

        _len_of_bag_of_words = _doc_sheet.max_row - 1

        for i in range(2, _len_of_bag_of_words + 2):

            word = _doc_sheet.cell(i, 1).value

            if _lemma_set.__contains__(word):
                tf = _lemma_set.get(word)
                _doc_sheet.cell(i, 58).value = tf
                idf = float(_doc_sheet.cell(i, 60).value)
                value = float(format(tf * idf, ".5f"))
                _doc_sheet.cell(i, 58).value = value
            else:
                _doc_sheet.cell(i, 58).value = 0

        print(datetime.now().strftime("%H:%M:%S") + ": completed query tf-idf calculations...")

        return

    def update_doc_sheet(self, _doc_sheet, _query):

        _lemma_set = self.process_query(_query)
        self.calculate_query_score(_doc_sheet, _lemma_set)
        _lemma_set.clear()

        return

    def pre_processing(self, stop_list_path, string):

        stop_list = self.import_stop_list(stop_list_path)

        # creating a object of the Preprocessing class
        _pre_processing = Preprocessing()

        # a stop-list is assigned to attribute of class
        _pre_processing.stop_word = stop_list

        # a string of file was passed to tokenizer that returns a list of tokens
        tokens = _pre_processing.tokenizer(string)

        # a set of lemma is returned as {lemma: tf}
        lemma_set = _pre_processing.lemmatizer(tokens)

        # list of stop-word is unusable, so it is cleared
        stop_list.clear()

        # list of tokens is unusable, so it is cleared
        tokens.clear()

        # returns a set of a doc: {lemma-0: tf, lemma-1: tf, ...}
        return lemma_set

    def prepare_bag_of_words(self, _lemmas_set, _bag_of_words):
        # lemma_set for i-th doc and keys() gives lemmas in the set
        lemmas = list(_lemmas_set.keys())
        for lem in lemmas:
            try:
                # a set was used because of its speed due to hashing which costs O(1)
                # otherwise a list would cost O(n) in searching the lemma from bag-of-words
                # because to restrict the repetitive lemmas from entire corpus
                _bag_of_words[lem] = 0
            except KeyError:
                # if lemma exists, it would throw KeyError and would do nothing
                pass
        # for loop ends
        return

    def calculate_doc_score(self, _bag_of_words, _len_of_bag_of_words, _lemmas, _sheet):

        # _bag_of_words = all the unique lemmas from the corpora
        # lemmas = [doc0={trump: 57, ...}, doc1={trump:1, ...}, ...]

        print(datetime.now().strftime("%H:%M:%S") + ": initiating document tf-idf calculations...")

        # iterates row-by-row on _sheet which is better than col-by-col on a _sheet of 6000 rows
        # _sheet is not 0 indexed, so its index starts from 1
        # the 1st row has column names from prepare_sheet()
        # from 2nd row we will start placing the words from bag-of-words
        # iterates until all the words from bag-of-words are placed in sheet
        for i in range(2, _len_of_bag_of_words + 2):

            # i-2 is because bag-of-words is 0 indexed, and this retrieves word from the bag
            # and places into sheet`s cell in 1st column and also assigning to word variable
            word = _sheet.cell(i, 1).value = _bag_of_words[i - 2]

            # 'df' is document-frequency: tells the no. of docs in which word has appeared
            df = 0

            # iterates over the all 56 docs from 0 to 55
            for doc_id in range(0, 56):
                # _lemmas[doc_id] gives a doc as set and get(word) gives the value the term-frequency
                # set was used because of hashing as retrieval is O(1)
                # otherwise list would cost O(n)
                tf = _lemmas[doc_id].get(word)

                if tf is None:
                    # tf is None when a word from bag does not appear in the doc
                    # tf will be 0 for this doc column in sheet
                    # doc_id + 2 is because doc_id starts from 0 and we have to skip the 'word' column
                    _sheet.cell(i, doc_id + 2).value = 0
                else:
                    # tf is assigned to that doc column
                    _sheet.cell(i, doc_id + 2).value = tf
                    # and df is incremented because word appears in this doc
                    df += 1
            # for loop ends

            # df is assigned to df column
            _sheet.cell(i, 59).value = df

            # idf = log10(N/df), N = size of corpora
            idf = float(format(math.log10(56 / df), '.5f'))
            _sheet.cell(i, 60).value = idf

            # tf`s are assigned to doc cols and idf is calculated
            # we iterate again over the doc cols and place tf*idf in doc cols
            for doc_id in range(0, 56):
                tf = float(_sheet.cell(i, doc_id + 2).value)
                value = float(format(tf * idf, '.5f'))
                _sheet.cell(i, doc_id + 2).value = value
            # for loop ends

        # for loop ends

        print(datetime.now().strftime("%H:%M:%S") + ": completed document tf-idf calculations")
        return _sheet

    def process_corpora(self, _bag_of_words, _lemmas):

        print(datetime.now().strftime("%H:%M:%S") + ": initiating pre-processing of documents...")

        # a loop is run to iterate over the entire corpus of length 56
        for i in range(0, 56):
            file_name = "../resource/trump-speeches/speech_" + str(i) + ".txt"
            file = open(file_name, "r")
            # skips the title
            file.readline()

            # lemma_set = {lemma: tf, ...}
            lemma_set = self.pre_processing("../resource/stopword-list.txt", file.read())

            # copy() deep copies the set into lemma list
            _lemmas.append(lemma_set.copy())

            # since tokens are created, file is now closed
            file.close()

            # since a deep-copy of lemma_set was appended and not its address was passed
            # this list can be cleared without effecting the one in lemmas
            lemma_set.clear()

            # _lemmas[i] gives a lemma_set for i-th doc
            # _bag_of_words and _lemmas[i] passed by reference
            self.prepare_bag_of_words(_lemmas[i], _bag_of_words)
        # for loop ends

        print(datetime.now().strftime("%H:%M:%S") + ": documents pre-processing completed")
        return

    def prepare_doc_sheet(self, _sheet):

        # total columns to be created: 1+56+1+1+1 = 60
        # <creating-columns>

        _sheet.cell(1, 1).value = "words"

        for i in range(2, 58):
            _sheet.cell(1, i).value = "d.tf:" + str(i - 2)
        # for loop ends

        _sheet.cell(1, 58).value = "q.tf"
        _sheet.cell(1, 59).value = "df"
        _sheet.cell(1, 60).value = "idf"

        # </creating-columns>
        return

    def create_doc_sheet(self, _sheet):

        print(datetime.now().strftime("%H:%M:%S") + ": preparing cache...")

        # _sheet is passed by reference
        self.prepare_doc_sheet(_sheet)

        # two important data-structures are initialized
        bag_of_words = {}
        lemmas = []

        # bag_of_words and lemmas are passed by reference
        self.process_corpora(bag_of_words, lemmas)

        # length of bag-of-words is saved for future use
        _len_of_bag_of_words = len(bag_of_words)

        print(datetime.now().strftime("%H:%M:%S") + ": length of bag-of-words: " + str(_len_of_bag_of_words))

        # keys from set 'bag-of-words' were converted into a list
        # so that keys() would not be called multiple times in the calculate_tf_idf
        bag_of_words = list(bag_of_words.keys())

        # a 'lemmas' list is in memory after being processed by 'process_corpora()'
        # which has sets for each document
        # each set has a key which is a lemma and a value against the key
        # as term-frequency
        # e.g: lemmas = [{trump: 57, ...}, {trump:1, ...}, ...]

        # bag_of_words, lemmas, and _sheet are passed by reference
        self.calculate_doc_score(bag_of_words, _len_of_bag_of_words, lemmas, _sheet)

        # since bag-of-words, and lemmas was written to disk
        # they are not needed in memory so they are cleared
        bag_of_words.clear()
        lemmas.clear()

        print(datetime.now().strftime("%H:%M:%S") + ": cache prepared")
        print(datetime.now().strftime("%H:%M:%S") + ": cleared not needed memory")

        # returns an excel sheet with all the doc vectors with scores
        return _sheet

    def create_result_set(self, _doc_sheet, alpha):

        print(datetime.now().strftime("%H:%M:%S") + ": creating result set...")

        _len_of_bag_of_words = _doc_sheet.max_row - 1
        _result_set = {}

        for doc_id in range(0, 56):

            scalar_product = 0
            norm_of_query_vector = 0.0
            norm_of_doc_vector = 0.0

            for i in range(2, _len_of_bag_of_words + 2):
                doc_tf_idf = float(_doc_sheet.cell(i, doc_id + 2).value)
                query_tf_idf = float(_doc_sheet.cell(i, 58).value)
                scalar_product += float(format(doc_tf_idf * query_tf_idf, ".5f"))
                norm_of_doc_vector += float(format(math.pow(doc_tf_idf, 2), '.5f'))
                norm_of_query_vector += float(format(math.pow(query_tf_idf, 2), '.5f'))

            norm_of_doc_vector = math.sqrt(norm_of_doc_vector)
            norm_of_query_vector = math.sqrt(norm_of_query_vector)

            # angle = cos0 = d . q / |d| . |q|

            angle = float(format(scalar_product / (norm_of_query_vector * norm_of_doc_vector), '.5f'))
            # angle = float(format(math.cos(product), ".5f"))
            if angle > alpha:
                _result_set[doc_id] = angle

        print(datetime.now().strftime("%H:%M:%S") + ": result set created")

        return _result_set

    def write_result_to_file(self, _path, _result_set, _query):

        print(datetime.now().strftime("%H:%M:%S") + ": writing result-set to " + _path + "...")

        file = open(_path, "a+")
        file.write("\n\nquery: " + _query + "\n")

        file.write("\nlength: " + str(len(_result_set)) + "\n")

        ranked_docs = []
        while bool(_result_set):
            ranked_docs.append(max(_result_set, key=_result_set.get))
            _result_set.pop(max(_result_set, key=_result_set.get))

        for doc in ranked_docs:
            file.write(str(doc) + ", ")

        file.close()
        print(datetime.now().strftime("%H:%M:%S") + ": result-set written to " + _path)
        return

    def import_nltk_data(self, path):
        # for lemmatizer, nltk_data was downloaded to resource
        # to use that data, we append its path to nltk`s data-path
        if not nltk.data.path.__contains__(path):
            print(datetime.now().strftime("%H:%M:%S") + ": appending nltk_data path to nltk`data-path...")
            nltk.data.path.append(path)
        return
