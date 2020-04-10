from datetime import datetime
import openpyxl
from src.VSM import VSM


# ---------------------------------- main code ----------------------------------
if __name__ == "__main__":

    # object of vsm class created that has methods for document and query processing
    vsm = VSM()
    vsm.stopword_file_path = "../resource/stopword-list.txt"

    # imports the data required for lemmatization
    vsm.import_nltk_data("../resource/nltk_data/")

    # 'workbook' is excel file object that contains a sheet which has bag-of-words
    # doc vectors, df, idf, and query vector
    workbook = 0
    # 'len_of_bag_of_words' has the total number of unique words from corpora
    # e.g: 6008 in my case
    len_of_bag_of_words = 0

    print(datetime.now().strftime("%H:%M:%S") + ": checking cache...")
    try:
        # if try successful, file is already created, no need to re-evaluate docs
        # load the workbook
        workbook = openpyxl.load_workbook("../out/tf-idf.xlsx")
        # load the sheet
        doc_sheet = workbook["Sheet"]
        print(datetime.now().strftime("%H:%M:%S") + ": cache found and loaded")
        len_of_bag_of_words = doc_sheet.max_row - 1
        print(datetime.now().strftime("%H:%M:%S") + ": length of bag-of-words = " + str(len_of_bag_of_words))

    except FileNotFoundError:
        print(datetime.now().strftime("%H:%M:%S") + ": cache not found")
        # file is not created, scores will be calculated, create a file/cache
        workbook = openpyxl.Workbook()
        # returns a sheet with doc-vectors with doc-tf-idf, df, idf and bag-of-words
        doc_sheet = vsm.create_doc_sheet(workbook.active)
        # saves the cache file on disk
        workbook.save("../out/tf-idf.xlsx")
        print(datetime.now().strftime("%H:%M:%S") + ": cache saved to disk")

    while True:

        print("*********************************")
        # takes query from user as string appends a space to the end of query for tokenizer handling
        query = input(datetime.now().strftime("%H:%M:%S") + ": search (0: exit): ") + " "
        # query = "pakistan afghanistan "

        if query == "0 ":
            break

        alpha = input(datetime.now().strftime("%H:%M:%S") + ": enter alpha: ")
        print("*********************************")
        # alpha = 0.0005

        if query != " " and alpha != "":

            alpha = float(alpha)
            # alpha = 0.0005

            # fills the query vector with tf-idf values
            vsm.update_doc_sheet(doc_sheet, query)
            # ret result-set as {doc-id: angle, doc-id: angle, ...}
            result_set = vsm.create_result_set(doc_sheet, alpha)
            vsm.write_result_to_file("../out/result_set.txt", result_set, query, alpha)
        else:
            print(datetime.now().strftime("%H:%M:%S") + ": empty query or alpha. try again !!!")

    # loop end

    workbook.close()
    print(datetime.now().strftime("%H:%M:%S") + ": cache-file closed")
    print(datetime.now().strftime("%H:%M:%S") + ": exit")
